"""Utility helpers for workbook parsing, normalization, and file output."""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
from datetime import date, datetime
import json
import logging
import math
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

LOGGER = logging.getLogger(__name__)

TOKEN_RE = re.compile(r"[^a-z0-9%]+")
PERCENT_RE = re.compile(r"(-?\d+(?:\.\d+)?)\s*%")
DATE_RE = re.compile(r"(\d{1,4}[/-]\d{1,2}[/-]\d{1,4})")


def safe_text(value: Any, default: str = "") -> str:
    """Convert workbook values to normalized text."""

    if value is None:
        return default
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return default
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value).strip() or default


def normalize_key(value: Any) -> str:
    """Create a matching key for fuzzy column detection."""

    text = safe_text(value).lower()
    text = text.replace("&", " and ")
    text = TOKEN_RE.sub(" ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_whitespace(value: Any) -> str:
    """Collapse extra whitespace in text values."""

    return re.sub(r"\s+", " ", safe_text(value)).strip()


def normalize_date(value: Any) -> str:
    """Normalize any date-like value to ISO format when possible."""

    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = safe_text(value)
    try:
        parsed = datetime.fromisoformat(text)
        return parsed.date().isoformat()
    except Exception:
        pass
    return text


def normalize_percentage(value: Any) -> float | None:
    """Convert percentage-like values to a fraction between 0 and 1."""

    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        numeric = float(value)
        if numeric > 1.0:
            numeric /= 100.0
        return max(0.0, min(1.0, numeric))
    text = safe_text(value)
    match = PERCENT_RE.search(text)
    if match:
        return max(0.0, min(1.0, float(match.group(1)) / 100.0))
    try:
        numeric = float(text)
        if numeric > 1.0:
            numeric /= 100.0
        return max(0.0, min(1.0, numeric))
    except ValueError:
        return None


def parse_number(value: Any) -> float | None:
    """Parse a number from a loosely formatted cell value."""

    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = safe_text(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def is_blank_row(values: Iterable[Any]) -> bool:
    """Determine whether a worksheet row is effectively empty."""

    return all(value in (None, "") for value in values)


def expand_merged_cells(sheet) -> None:
    """Propagate merged-cell values into the full merged range."""

    for merged_range in list(sheet.merged_cells.ranges):
        top_left = sheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in sheet.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    sheet[cell.coordinate] = top_left


def load_workbook_safely(path: Path):
    """Load an Excel workbook without crashing on corrupt files."""

    try:
        return load_workbook(path, data_only=False)
    except Exception as exc:
        LOGGER.exception("Failed to load workbook %s", path)
        raise RuntimeError(f"Unable to read workbook {path.name}: {exc}") from exc


def deduplicate_dicts(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Remove duplicate dictionaries while preserving order."""

    seen: set[str] = set()
    unique_records: list[dict[str, Any]] = []
    for record in records:
        signature = json.dumps(record, sort_keys=True, default=str)
        if signature in seen:
            continue
        seen.add(signature)
        unique_records.append(record)
    return unique_records


def to_json_serializable(value: Any) -> Any:
    """Convert dataclasses and Path values into JSON-safe structures."""

    if is_dataclass(value):
        return asdict(value)
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, list):
        return [to_json_serializable(item) for item in value]
    if isinstance(value, dict):
        return {key: to_json_serializable(item) for key, item in value.items()}
    return value


def write_text_file(path: Path, content: str) -> None:
    """Write UTF-8 text safely, creating parent directories if needed."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def ensure_directory(path: Path) -> Path:
    """Create a directory if it does not already exist."""

    path.mkdir(parents=True, exist_ok=True)
    return path


def extract_first_date(text: str) -> str:
    """Return the first date-like token from text if present."""

    match = DATE_RE.search(text)
    return match.group(1) if match else ""


def slugify(value: str) -> str:
    """Convert arbitrary text into a filesystem-safe slug."""

    cleaned = normalize_key(value)
    cleaned = cleaned.replace(" ", "_")
    return cleaned or "item"
